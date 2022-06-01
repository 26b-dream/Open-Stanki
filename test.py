from __future__ import annotations
import copy
from time import sleep

from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from playwright.sync_api._generated import Response

import re
from bs4 import BeautifulSoup
from common.constants import BASE_DIR, DOWNLOADED_FILES_DIR
from common.extended_path import ExtendedPath
from common.extended_playwright import sync_playwright
from playwright.sync_api._generated import BrowserContext
from openpyxl import load_workbook


class Oxford:
    BASE_URL = "https://learninglink.oup.com"
    FILES_DIR = DOWNLOADED_FILES_DIR / "Oxford"

    def __init__(self, url: str) -> None:
        self.main_url = url
        self.name = str(ExtendedPath.convert_to_path(ExtendedPath(url).name))

    def import_all(self) -> None:
        with sync_playwright() as p:
            browser = p.chromium.launch_persistent_context(
                DOWNLOADED_FILES_DIR / "cookies/Chrome",
                headless=False,
                accept_downloads=True,
                channel="chrome",
            )

            main_page_path = self.FILES_DIR / ExtendedPath(self.main_url).remove_parent(1).with_suffix(".html")
            # Download main page if needed
            if not main_page_path.exists():
                # Open main page
                page = browser.new_page()
                page.goto(self.main_url, wait_until="networkidle")
                main_page_path.write(page.content())
                page.close()

            parsed_html = main_page_path.parse_html()
            self.book_title = parsed_html.strict_select_one("title").text

            self.download_flashcards(browser, parsed_html)
            self.import_flashcards(parsed_html)

            self.download_quiz(browser, parsed_html)
            self.import_quiz(parsed_html)

    def download_quiz(self, browser: BrowserContext, parsed_html: BeautifulSoup) -> None:
        for article in parsed_html.strict_select("article"):
            title = article.get("title")

            # Ignore articles with no titles
            # Ignore articles that return a list because I have no idea what those are
            if title is None or isinstance(title, list):
                continue

            if title.endswith("Quiz Without Consequence") or title.endswith("Pre-Test"):
                url = self.BASE_URL + article.strict_get_str("data-linktarget")
                partial_path = ExtendedPath(url).remove_parent(1).legalize().with_suffix(".html")
                html_path = self.FILES_DIR / partial_path

                # Download Quiz page if needed
                if not html_path.exists():
                    page = browser.new_page()
                    page.goto(url, wait_until="networkidle")
                    page.click("button[type='submit']")
                    page.click("button[id='ckSubmit']")
                    # waitforloadstate doesn't work so just throw a sleep here
                    sleep(1)
                    html_path.write(page.content())
                    page.close()

    def import_quiz(self, parsed_html: BeautifulSoup) -> None:
        for article in parsed_html.strict_select("article"):
            title = article.get("title")

            # Ignore articles with no titles
            # Ignore articles that return a list because I have no idea what those are
            if title is None or isinstance(title, list):
                continue

            if title.endswith("Quiz Without Consequence") or title.endswith("Pre-Test"):
                output: list[list[str]] = []
                quiz_url = self.BASE_URL + article.strict_get_str("data-linktarget")
                quiz_partial_path = ExtendedPath(quiz_url).remove_parent(1).legalize().with_suffix(".html")
                quiz_html_path = self.FILES_DIR / quiz_partial_path
                parsed_quiz = quiz_html_path.parse_html()

                # Go through each question
                for i, section in enumerate(parsed_quiz.strict_select("section[class^='indQuest']")):
                    question = section.strict_select_one("span[class='txt']").text

                    # Create answer
                    question_html = f"<b>{question}</b><br>"
                    for span in section.strict_select("label"):
                        raw_question_option = span.strict_select("span")[0].text
                        question_html += f"<input type='radio'>{raw_question_option}<br>"

                    # Create answer
                    answer_html = f"<b>{question}</b><br>"
                    for label in section.strict_select("label"):
                        spans = label.strict_select("span")
                        raw_question_option = spans[0].text
                        span_class = str(spans[1].strict_get_str("class"))

                        # If the span is labeled as d-none the text is hidden so it is the wrong answer
                        if "d-none" in span_class:
                            answer_html += "❌"
                        else:
                            answer_html += "✅"
                        answer_html += f"{raw_question_option}<br>"

                    id = " - ".join(
                        ["Oxford", self.book_title, parsed_quiz.strict_select_one("title").text, str(i + 1)]
                    )
                    output.append([id, question_html, answer_html])

                output_path = (
                    BASE_DIR / "Output" / "Oxford" / self.book_title / title.removeprefix("View ")
                ).with_suffix(".csv")
                output_path.write_csv(output)

    def download_flashcards(self, browser: BrowserContext, parsed_html: BeautifulSoup) -> None:
        def download_xlsx(response: Response):
            if response.url.endswith(".xlsx"):
                partial_xlsx_path = ExtendedPath(response.url).remove_parent()
                xlsx_path = self.FILES_DIR / partial_xlsx_path
                xlsx_path.write(response.body())

        for article in parsed_html.strict_select("article"):
            title = article.get("title")

            # Ignore articles with no titles
            # Ignore articles that return a list because I have no idea what those are
            if title is None or isinstance(title, list):
                continue

            if title.endswith("Flashcards"):
                url = self.BASE_URL + article.strict_get_str("data-linktarget")

                partial_path = ExtendedPath(url).remove_parent(1).legalize().with_suffix(".html")
                html_path = self.FILES_DIR / partial_path

                # Download flashcard page if needed
                if not html_path.exists():
                    page = browser.new_page()
                    page.on("response", download_xlsx)
                    page.goto(url, wait_until="networkidle")
                    html_path.write(page.content())
                    page.close()

    def import_flashcards(self, parsed_html: BeautifulSoup) -> None:
        for article in parsed_html.strict_select("article"):
            title = article.get("title")

            # Ignore articles with no titles
            # Ignore articles that return a list because I have no idea what those are
            if title is None or isinstance(title, list):
                continue

            if title.endswith("Flashcards"):
                flashcard_url = self.BASE_URL + article.strict_get_str("data-linktarget")
                partial_flashcard_path = ExtendedPath(flashcard_url).remove_parent(1).legalize().with_suffix(".html")
                flashcard_html_path = self.FILES_DIR / partial_flashcard_path
                parsed_flashcard_html = flashcard_html_path.parse_html()
                javascript_content = parsed_flashcard_html.strict_select("script[language='JavaScript']")[0].text
                partial_xlsx_url = re.search(r"var isXlsx=true, dataFile = \"/(.*?)\"", javascript_content).group(1)
                xlsx_path = self.FILES_DIR / partial_xlsx_url
                parsed_xlsx = load_workbook(xlsx_path)
                sheet = parsed_xlsx[parsed_xlsx.sheetnames[0]]

                output: list[list[str]] = []
                for i, cell in enumerate(zip(sheet["A"], sheet["B"])):
                    word = cell[0].value
                    definition = cell[1].value
                    id = " - ".join(
                        ["Oxford", self.book_title, parsed_flashcard_html.strict_select_one("title").text, word]
                    )

                    output.append([id, word, definition])
                output_path = (
                    BASE_DIR / "Output" / "Oxford" / self.book_title / title.removeprefix("View ")
                ).with_suffix(".csv")
                output_path.write_csv(output)


# <script language="JavaScript">
# 	//find the file path for the csv
# 	var isXlsx=true, dataFile = "/protected/files/content/flashcardXlsx/1573540802783-Ch01.xlsx";

# 	init(isXlsx,dataFile);
# </script>


# <a href="#" id="button_next" tabindex="0" role="button" aria-label="Next Card" class="btn btn-block g-color-white g-color-white--hover g-bg-primary g-bg-main--hover g-rounded-30 mb-2" data-toggle="tooltip" title="" data-placement="left" data-original-title="Next Card">
#                 <i class="fa fa-mail-forward mobile-only"></i>
#                 <span class="g-hidden-sm-down" aria-hidden="true" id="nextTxt">Next Card</span>
#               </a>

# Read each line from urls.txt
for line in open("urls.txt"):
    # Clean URL
    line = line.strip()

    Oxford(line).import_all()
